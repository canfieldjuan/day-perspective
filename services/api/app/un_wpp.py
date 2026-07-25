from __future__ import annotations

import calendar
import csv
import hashlib
import io
import urllib.request
from dataclasses import dataclass
from datetime import UTC, date, datetime
from decimal import ROUND_HALF_UP, Decimal
from io import BytesIO
from pathlib import Path
from typing import Any
from uuid import UUID

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.adapters.base import LocalFilesystemRawSourceStore, RawSourceStore
from app.governance import (
    EditorialSelectionStatus,
    LicenseInput,
    ReviewDecisionValue,
    assert_release_publication_eligible,
    record_claim_review,
    record_editorial_selection,
    register_release_license,
)
from app.models import (
    Claim,
    ClaimAssertionStatus,
    ComparabilityStatus,
    DataStatus,
    DateRole,
    DerivedValue,
    DerivedValueInput,
    LegalReviewStatus,
    Methodology,
    Metric,
    MetricCoverage,
    Observation,
    PipelineRun,
    QualityAssessment,
    QualityCheck,
    RawSourceRecord,
    ResolutionMethod,
    ResolvedClaim,
    ResolvedClaimEvidence,
    ReviewTask,
    Source,
    SourceRelease,
    TemporalAssignment,
    TemporalPrecision,
)
from app.services import (
    PublicationStatementEvidenceInput,
    canonical_json_bytes,
    content_hash,
    create_claim,
    create_source_release,
    resolve_claim,
)

__all__ = ["LocalFilesystemRawSourceStore"]

UN_WPP_SOURCE_SLUG = "un-wpp-2024"
UN_WPP_SOURCE_URL = (
    "https://population.un.org/wpp/assets/Excel%20Files/"
    "1_Indicator%20(Standard)/EXCEL_FILES/1_General/"
    "WPP2024_GEN_F01_DEMOGRAPHIC_INDICATORS_COMPACT.xlsx"
)
UN_WPP_TERMS_URL = "https://creativecommons.org/licenses/by/3.0/igo/"
UN_WPP_LICENSE_SNAPSHOT = (
    "The official WPP 2024 GEN/01/REV1 workbook states that the United Nations "
    "work is available under CC BY 3.0 IGO with a suggested source citation."
)
UN_WPP_WORKBOOK_SHA256 = (
    "98e34d9b65b53858cd08a57a566e45050b08093ad85ba5714fe6fbd78055ae6d"
)
UN_WPP_NORMALIZED_FIXTURE_SHA256 = (
    "a491914da7e942feb016304ada0c9c43f5c7d576fc15bf3541ba54c16a4aa039"
)
GOLDEN_YEAR = 1964
GOLDEN_DATE = date(1964, 3, 27)
SUPPORTED_YEARS = frozenset(range(1950, 2026))
ESTIMATE_YEARS = frozenset(range(1950, 2024))
PROJECTION_YEARS = frozenset({2024, 2025})
BASELINE_YEAR = min(SUPPORTED_YEARS)

METRIC_DEFINITIONS = {
    "population_midyear": (
        "World mid-year population",
        "thousand persons",
        "Estimated population at 1 July of the calendar year.",
    ),
    "annual_births": (
        "World annual births",
        "thousand persons",
        "Estimated live births during the calendar year.",
    ),
    "annual_deaths": (
        "World annual deaths",
        "thousand persons",
        "Estimated deaths during the calendar year.",
    ),
    "life_expectancy": (
        "World life expectancy at birth",
        "years",
        "Estimated period life expectancy at birth for the calendar year.",
    ),
    "under_five_mortality": (
        "World under-five mortality",
        "deaths per 1,000 live births",
        "Estimated probability of dying before age five per 1,000 live births.",
    ),
}


@dataclass(frozen=True)
class WPPRecord:
    location: str
    location_code: str
    year: int
    variant: str
    population_july_thousands: Decimal
    births_thousands: Decimal
    deaths_thousands: Decimal
    life_expectancy_years: Decimal
    under_five_mortality_per_1000: Decimal

    @property
    def record_id(self) -> str:
        return f"{self.location_code}:{self.year}:{self.variant.lower()}"

    @property
    def data_status(self) -> DataStatus:
        return (
            DataStatus.ESTIMATED
            if self.variant == "Estimates"
            else DataStatus.MODELED
        )

    def canonical_value(self) -> dict[str, str | int]:
        return {
            "location": self.location,
            "location_code": self.location_code,
            "year": self.year,
            "variant": self.variant,
            "data_status": self.data_status.value,
            "population_july_thousands": str(self.population_july_thousands),
            "births_thousands": str(self.births_thousands),
            "deaths_thousands": str(self.deaths_thousands),
            "life_expectancy_years": str(self.life_expectancy_years),
            "under_five_mortality_per_1000": str(
                self.under_five_mortality_per_1000
            ),
        }


@dataclass(frozen=True)
class WPPIngestionResult:
    pipeline_run_id: UUID
    source_release_id: UUID | None
    claim_count: int
    checksum: str
    idempotent: bool


@dataclass(frozen=True)
class WPPProfileContent:
    typical_statements: list[dict[str, object]]
    context_statements: list[dict[str, object]]
    evidence: list[PublicationStatementEvidenceInput]
    source_release_id: UUID
    methodology: Methodology
    resolved_claims: tuple[ResolvedClaim, ...]


NORMALIZED_FIELDS = {
    "location",
    "location_code",
    "year",
    "variant",
    "population_july_thousands",
    "births_thousands",
    "deaths_thousands",
    "life_expectancy_years",
    "under_five_mortality_per_1000",
}

WORKBOOK_COLUMNS = {
    "variant": (1, "Variant"),
    "location": (2, "Region, subregion, country or area *"),
    "location_code": (4, "Location code"),
    "year": (10, "Year"),
    "population_july_thousands": (
        12,
        "Total Population, as of 1 July (thousands)",
    ),
    "births_thousands": (23, "Births (thousands)"),
    "deaths_thousands": (30, "Total Deaths (thousands)"),
    "life_expectancy_years": (
        34,
        "Life Expectancy at Birth, both sexes (years)",
    ),
    "under_five_mortality_per_1000": (
        50,
        "Under-Five Mortality (deaths under age 5 per 1,000 live births)",
    ),
}


def _record(
    *,
    location: object,
    location_code: object,
    year: object,
    variant: object,
    population_july_thousands: object,
    births_thousands: object,
    deaths_thousands: object,
    life_expectancy_years: object,
    under_five_mortality_per_1000: object,
) -> WPPRecord:
    try:
        return WPPRecord(
            location=str(location),
            location_code=str(location_code),
            year=int(str(year)),
            variant=str(variant),
            population_july_thousands=Decimal(
                str(population_july_thousands)
            ),
            births_thousands=Decimal(str(births_thousands)),
            deaths_thousands=Decimal(str(deaths_thousands)),
            life_expectancy_years=Decimal(str(life_expectancy_years)),
            under_five_mortality_per_1000=Decimal(
                str(under_five_mortality_per_1000)
            ),
        )
    except (TypeError, ValueError, ArithmeticError) as error:
        raise ValueError("The WPP release contains an invalid measure.") from error


def _validate_records(records: list[WPPRecord]) -> tuple[WPPRecord, ...]:
    if not records or len({record.record_id for record in records}) != len(
        records
    ):
        raise ValueError("The WPP release must contain unique World-year records.")
    for record in records:
        expected_variant = "Estimates" if record.year in ESTIMATE_YEARS else "Medium"
        if (
            record.location != "World"
            or record.location_code != "900"
            or record.year not in SUPPORTED_YEARS
            or record.variant != expected_variant
        ):
            raise ValueError("The WPP release contains an unsupported record.")
        measures = (
            record.population_july_thousands,
            record.births_thousands,
            record.deaths_thousands,
            record.life_expectancy_years,
            record.under_five_mortality_per_1000,
        )
        if not all(value.is_finite() for value in measures):
            raise ValueError("The WPP release contains a non-finite measure.")
        if min(measures) <= 0:
            raise ValueError("The WPP release contains a non-positive measure.")
    years = {record.year for record in records}
    if years != SUPPORTED_YEARS or len(records) != len(SUPPORTED_YEARS):
        raise ValueError(
            "The WPP release must contain exactly one World record for every "
            "year from 1950 through 2025."
        )
    return tuple(sorted(records, key=lambda record: record.year))


def _parse_normalized_csv(payload: bytes) -> tuple[WPPRecord, ...]:
    reader = csv.DictReader(io.StringIO(payload.decode("utf-8")))
    if reader.fieldnames is None or set(reader.fieldnames) != NORMALIZED_FIELDS:
        raise ValueError("The WPP normalized schema does not match GEN/01/REV1.")
    records: list[WPPRecord] = []
    for row in reader:
        records.append(
            _record(
                location=row["location"],
                location_code=row["location_code"],
                year=row["year"],
                variant=row["variant"],
                population_july_thousands=row["population_july_thousands"],
                births_thousands=row["births_thousands"],
                deaths_thousands=row["deaths_thousands"],
                life_expectancy_years=row["life_expectancy_years"],
                under_five_mortality_per_1000=row[
                    "under_five_mortality_per_1000"
                ],
            )
        )
    return _validate_records(records)


def _parse_workbook(payload: bytes) -> tuple[WPPRecord, ...]:
    workbook = load_workbook(
        filename=BytesIO(payload),
        read_only=True,
        data_only=True,
    )
    required_sheets = {"Estimates", "Medium variant"}
    if not required_sheets <= set(workbook.sheetnames):
        raise ValueError("The WPP workbook is missing a required sheet.")
    records: list[WPPRecord] = []
    sheet_specs = (
        ("Estimates", ESTIMATE_YEARS),
        ("Medium variant", PROJECTION_YEARS),
    )
    for sheet_name, supported_years in sheet_specs:
        sheet = workbook[sheet_name]
        header = next(
            sheet.iter_rows(min_row=17, max_row=17, values_only=True)
        )
        for index, expected in WORKBOOK_COLUMNS.values():
            if header[index] != expected:
                raise ValueError(
                    f"The WPP workbook column contract changed on {sheet_name}."
                )
        found_world = False
        for row in sheet.iter_rows(min_row=18, values_only=True):
            location_code = row[WORKBOOK_COLUMNS["location_code"][0]]
            if str(location_code) not in {"900", "900.0"}:
                if found_world:
                    break
                continue
            found_world = True
            year = int(str(row[WORKBOOK_COLUMNS["year"][0]]))
            if year not in supported_years:
                continue
            values: dict[str, Any] = {
                name: row[index] for name, (index, _) in WORKBOOK_COLUMNS.items()
            }
            records.append(_record(**values))
    workbook.close()
    return _validate_records(records)


def _parse(payload: bytes, *, input_format: str) -> tuple[WPPRecord, ...]:
    if input_format == "normalized_csv":
        return _parse_normalized_csv(payload)
    if input_format == "official_xlsx":
        return _parse_workbook(payload)
    raise ValueError(f"Unsupported WPP input format: {input_format}.")


def _retrieve(fixture_path: Path | None) -> tuple[bytes, str]:
    if fixture_path is not None:
        return (
            fixture_path.read_bytes(),
            "official_xlsx"
            if fixture_path.suffix.lower() == ".xlsx"
            else "normalized_csv",
        )
    request = urllib.request.Request(
        UN_WPP_SOURCE_URL,
        headers={"User-Agent": "day-perspective-offline-ingestion/0.4"},
    )
    with urllib.request.urlopen(request, timeout=120) as response:
        return response.read(), "official_xlsx"


def _methodology(session: Session) -> Methodology:
    existing = session.scalar(
        select(Methodology).where(
            Methodology.slug == "un-wpp-annual-context",
            Methodology.version == "2",
        )
    )
    if existing is not None:
        return existing
    definition = {
        "source": "UN WPP 2024 estimates",
        "daily_equivalent": (
            "annual total in persons divided by the Gregorian day count for the year"
        ),
        "language": (
            "Average daily equivalent based on annual total; never a date observation."
        ),
    }
    row = Methodology(
        slug="un-wpp-annual-context",
        version="2",
        name="UN WPP annual context and daily-equivalent method",
        description=definition["language"],
        method_kind="annual_context_and_uniform_allocation",
        formula="annual_total_persons / gregorian_days_in_year",
        code_version="0.4.0",
        definition_hash=content_hash(definition),
        legal_review_status=LegalReviewStatus.NOT_REQUIRED,
    )
    session.add(row)
    session.flush()
    return row


def _register_license(session: Session, release_id: UUID) -> None:
    register_release_license(
        session,
        source_release_id=release_id,
        license_input=LicenseInput(
            license_identifier="CC-BY-3.0-IGO",
            license_snapshot=UN_WPP_LICENSE_SNAPSHOT,
            terms_url=UN_WPP_TERMS_URL,
            commercial_use_permission=True,
            redistribution_permission=True,
            derivatives_permission=True,
            attribution_required=True,
            attribution_text=(
                "United Nations DESA Population Division (2024), "
                "World Population Prospects 2024, CC BY 3.0 IGO."
            ),
            public_display_permission=True,
            raw_download_permission=True,
            terms_checked_at=date(2026, 7, 24),
            legal_review_status=LegalReviewStatus.NOT_REQUIRED,
        ),
    )


def ingest_un_wpp(
    session: Session,
    *,
    fixture_path: Path | None = None,
    raw_store: RawSourceStore,
    dry_run: bool = False,
) -> WPPIngestionResult:
    mode = "fixture" if fixture_path is not None else "live"
    input_format = (
        "official_xlsx"
        if fixture_path is None or fixture_path.suffix.lower() == ".xlsx"
        else "normalized_csv"
    )
    run = PipelineRun(
        pipeline_name="un-wpp-2024-adapter",
        code_version="0.4.0",
        configuration_hash=content_hash(
            {
                "dataset": "GEN/01/REV1",
                "mode": mode,
                "input_format": input_format,
                "source_url": UN_WPP_SOURCE_URL,
            }
        ),
        status="running",
        details={"mode": mode, "input_format": input_format},
    )
    session.add(run)
    session.flush()
    stage = "retrieval"
    try:
        payload, retrieved_format = _retrieve(fixture_path)
        if retrieved_format != input_format:
            raise ValueError(
                f"Expected {input_format} input but retrieved {retrieved_format}."
            )
        checksum = hashlib.sha256(payload).hexdigest()
        stage = "validation"
        if mode == "live" and checksum != UN_WPP_WORKBOOK_SHA256:
            raise ValueError(
                "Live UN WPP workbook checksum does not match pinned GEN/01/REV1."
            )
        records = _parse(payload, input_format=input_format)
        if dry_run:
            run.status = "succeeded"
            run.completed_at = datetime.now(UTC)
            run.details = {
                **run.details,
                "dry_run": True,
                "checksum": checksum,
                "records": len(records),
            }
            session.add(
                QualityCheck(
                    pipeline_run_id=run.id,
                    check_name="un_wpp_schema_supported_world_years",
                    status="passed",
                    subject_type="pipeline_run",
                    subject_id=run.id,
                    details={
                        "dry_run": True,
                        "years": [record.year for record in records],
                    },
                )
            )
            return WPPIngestionResult(run.id, None, 0, checksum, False)
        stage = "persistence"
        with session.begin_nested():
            source = session.scalar(
                select(Source).where(Source.slug == UN_WPP_SOURCE_SLUG)
            )
            if source is None:
                source = Source(
                    slug=UN_WPP_SOURCE_SLUG,
                    name="World Population Prospects 2024",
                    publisher=(
                        "United Nations, Department of Economic and Social Affairs, "
                        "Population Division"
                    ),
                    canonical_url="https://population.un.org/wpp/",
                    legal_review_status=LegalReviewStatus.NOT_REQUIRED,
                )
                session.add(source)
                session.flush()
            else:
                source.name = "World Population Prospects 2024"
                source.publisher = (
                    "United Nations, Department of Economic and Social Affairs, "
                    "Population Division"
                )
                source.canonical_url = "https://population.un.org/wpp/"
            existing = session.scalar(
                select(SourceRelease).where(
                    SourceRelease.source_id == source.id,
                    SourceRelease.raw_checksum_sha256 == checksum,
                )
            )
            if existing is not None:
                raw_store.read(
                    existing.raw_storage_uri,
                    existing.raw_checksum_sha256,
                )
                _register_license(session, existing.id)
                run.status = "succeeded"
                run.completed_at = datetime.now(UTC)
                run.details = {**run.details, "idempotent": True}
                claim_count = len(
                    list(
                        session.scalars(
                            select(Claim.id).where(
                                Claim.source_release_id == existing.id
                            )
                        )
                    )
                )
                return WPPIngestionResult(
                    run.id, existing.id, claim_count, checksum, True
                )
            storage_uri = raw_store.write(UN_WPP_SOURCE_SLUG, checksum, payload)
            release = create_source_release(
                session,
                source_id=source.id,
                release_label=f"wpp2024-gen01-rev1-{checksum[:12]}",
                source_url=UN_WPP_SOURCE_URL,
                raw_storage_uri=storage_uri,
                raw_bytes=payload,
                raw_record_count=len(records),
                pipeline_run_id=run.id,
                metadata_json={
                    "dataset": "World Population Prospects 2024",
                    "quality_contract_version": "1",
                    "required_quality_checks": [
                        "un_wpp_schema_supported_world_years"
                    ],
                    "file_identity": "GEN/01/REV1",
                    "input_format": input_format,
                    "fixture": fixture_path is not None,
                    **(
                        {"upstream_source_file_sha256": checksum}
                        if input_format == "official_xlsx"
                        else {
                            "upstream_source_file_sha256": UN_WPP_WORKBOOK_SHA256
                        }
                        if checksum == UN_WPP_NORMALIZED_FIXTURE_SHA256
                        else {}
                    ),
                    "variants": ["Estimates", "Medium"],
                    "supported_year_start": min(SUPPORTED_YEARS),
                    "supported_year_end": max(SUPPORTED_YEARS),
                    "license": "CC-BY-3.0-IGO",
                },
                legal_review_status=LegalReviewStatus.NOT_REQUIRED,
            )
            _register_license(session, release.id)
            claim_count = 0
            metric_specs = (
                ("population_midyear", "population_july_thousands", "thousand persons"),
                ("annual_births", "births_thousands", "thousand persons"),
                ("annual_deaths", "deaths_thousands", "thousand persons"),
                ("life_expectancy", "life_expectancy_years", "years"),
                (
                    "under_five_mortality",
                    "under_five_mortality_per_1000",
                    "deaths per 1,000 live births",
                ),
            )
            for record in records:
                record_json = record.canonical_value()
                record_hash = hashlib.sha256(
                    canonical_json_bytes(record_json)
                ).hexdigest()
                locator = f"{UN_WPP_SOURCE_URL}#World-{record.year}"
                session.add(
                    RawSourceRecord(
                        source_release_id=release.id,
                        source_record_id=record.record_id,
                        source_record_locator=locator,
                        raw_storage_uri=storage_uri,
                        raw_checksum_sha256=record_hash,
                        schema_version="wpp2024-gen01-supported-years-v1",
                        payload_json=record_json,
                    )
                )
                for predicate, attribute, unit in metric_specs:
                    value = getattr(record, attribute)
                    claim = create_claim(
                        session,
                        source_release_id=release.id,
                        source_record_locator=locator,
                        source_record_hash_sha256=record_hash,
                        claim_type=predicate,
                        assertion_text=f"{value} {unit}",
                        assertion_json={
                            "value": str(value),
                            "unit": unit,
                            "year": record.year,
                            "geography": "World",
                            "variant": record.variant,
                            "data_status": record.data_status.value,
                        },
                        assertion_status=ClaimAssertionStatus.CANDIDATE,
                        unit=unit,
                        lower_bound=value,
                        upper_bound=value,
                    )
                    claim.temporal_start = date(record.year, 1, 1)
                    claim.temporal_end = date(record.year, 12, 31)
                    claim.temporal_precision = TemporalPrecision.YEAR
                    claim.temporal_assignment = TemporalAssignment.DIRECT_RECORD
                    claim.date_role = DateRole.REPORTED
                    claim.data_status = record.data_status
                    claim.pipeline_run_id = run.id
                    session.add(
                        ReviewTask(
                            claim_id=claim.id,
                            status="open",
                            priority="normal",
                            rationale=(
                                f"Review UN WPP {predicate} estimate for {record.year}."
                            ),
                        )
                    )
                    claim_count += 1
            session.add(
                QualityCheck(
                    pipeline_run_id=run.id,
                    check_name="un_wpp_schema_supported_world_years",
                    status="passed",
                    subject_type="source_release",
                    subject_id=release.id,
                    details={
                        "records": len(records),
                        "claims": claim_count,
                        "years": [record.year for record in records],
                        "estimate_years": [
                            record.year
                            for record in records
                            if record.data_status == DataStatus.ESTIMATED
                        ],
                        "projection_years": [
                            record.year
                            for record in records
                            if record.data_status == DataStatus.MODELED
                        ],
                    },
                )
            )
            run.status = "succeeded"
            run.completed_at = datetime.now(UTC)
            run.details = {**run.details, "idempotent": False, "checksum": checksum}
            return WPPIngestionResult(
                run.id, release.id, claim_count, checksum, False
            )
    except Exception as error:
        run.status = "failed"
        run.completed_at = datetime.now(UTC)
        run.details = {
            **run.details,
            "error": type(error).__name__,
            "failure_stage": stage,
        }
        check_name = {
            "retrieval": "un_wpp_retrieval",
            "validation": "un_wpp_schema_supported_world_years",
            "persistence": "un_wpp_ingestion",
        }[stage]
        session.add(
            QualityCheck(
                pipeline_run_id=run.id,
                check_name=check_name,
                status="failed",
                subject_type="pipeline_run",
                subject_id=run.id,
                details={"error": str(error), "failure_stage": stage},
            )
        )
        session.flush()
        raise


def review_un_wpp(
    session: Session,
    source_release_id: UUID,
    *,
    editorial_dates: tuple[date, ...] = (GOLDEN_DATE,),
) -> int:
    claims = list(
        session.scalars(
            select(Claim)
            .where(Claim.source_release_id == source_release_id)
            .order_by(Claim.claim_type, Claim.temporal_start)
        )
    )
    required_predicates = set(METRIC_DEFINITIONS)
    claims_by_year: dict[int, list[Claim]] = {}
    for claim in claims:
        if claim.temporal_start is None:
            raise ValueError("Every WPP claim requires an annual period.")
        claims_by_year.setdefault(claim.temporal_start.year, []).append(claim)
    if set(claims_by_year) != SUPPORTED_YEARS or any(
        {claim.claim_type for claim in year_claims} != required_predicates
        or len(year_claims) != len(required_predicates)
        for year_claims in claims_by_year.values()
    ):
        raise ValueError(
            "WPP review requires five claims for every year from 1950 through 2025."
        )
    if any(selected.year not in SUPPORTED_YEARS for selected in editorial_dates):
        raise ValueError("WPP editorial dates must use a supported year.")
    blocked = [
        claim.claim_type
        for claim in claims
        if claim.assertion_status
        not in {
            ClaimAssertionStatus.CANDIDATE,
            ClaimAssertionStatus.IN_REVIEW,
            ClaimAssertionStatus.ACCEPTED,
        }
    ]
    if blocked:
        raise ValueError(
            "Non-accepted WPP claims block review: " + ", ".join(sorted(blocked))
        )
    methodology = _methodology(session)
    resolved_count = 0
    for claim in claims:
        if claim.assertion_status in {
            ClaimAssertionStatus.CANDIDATE,
            ClaimAssertionStatus.IN_REVIEW,
        }:
            record_claim_review(
                session,
                claim=claim,
                decision=ReviewDecisionValue.ACCEPTED,
                rationale="Matched the selected row and metric to WPP GEN/01/REV1.",
                reviewed_by="development-fixture-review",
            )
        if claim.assertion_status != ClaimAssertionStatus.ACCEPTED:
            raise ValueError("Every WPP claim must be accepted before resolution.")
        year = claim.temporal_start.year if claim.temporal_start is not None else 0
        canonical_key = f"un-wpp:world:{year}:{claim.claim_type}"
        prior = session.scalar(
            select(ResolvedClaim)
            .where(ResolvedClaim.canonical_key == canonical_key)
            .order_by(ResolvedClaim.version.desc())
        )
        prior_supports_current = (
            prior is not None
            and session.scalar(
                select(ResolvedClaimEvidence.claim_id).where(
                    ResolvedClaimEvidence.resolved_claim_id == prior.id,
                    ResolvedClaimEvidence.claim_id == claim.id,
                    ResolvedClaimEvidence.stance == "supporting",
                )
            )
            is not None
        )
        if (
            prior is None
            or prior.resolved_value != (claim.assertion_json or {})
            or not prior_supports_current
        ):
            resolved_row = resolve_claim(
                session,
                canonical_key=canonical_key,
                resolved_value=claim.assertion_json or {},
                rationale=(
                    "Accepted one attributed official UN WPP "
                    + (
                        "medium-variant projection. The modeled"
                        if claim.data_status == DataStatus.MODELED
                        else "estimate. The estimated"
                    )
                    + " status and annual resolution remain visible."
                ),
                supporting_claim_ids=[claim.id],
                resolution_method=ResolutionMethod.SINGLE_SOURCE,
                methodology_id=methodology.id,
                supersedes_resolved_claim_id=prior.id if prior is not None else None,
            )
            resolved_row.comparability_status = ComparabilityStatus.COMPARABLE
        else:
            resolved_row = prior
        display_name, unit, definition = METRIC_DEFINITIONS[claim.claim_type]
        metric = session.scalar(
            select(Metric).where(Metric.metric_key == f"un-wpp:{claim.claim_type}")
        )
        if metric is None:
            metric = Metric(
                metric_key=f"un-wpp:{claim.claim_type}",
                display_name=display_name,
                unit=unit,
                definition=definition,
                data_status=DataStatus.ESTIMATED,
                provenance_resolved_claim_id=resolved_row.id,
                methodology_id=methodology.id,
            )
            session.add(metric)
            session.flush()
        elif year == BASELINE_YEAR:
            metric.provenance_resolved_claim_id = resolved_row.id
            metric.methodology_id = methodology.id
        observation = session.scalar(
            select(Observation).where(
                Observation.metric_id == metric.id,
                Observation.source_release_id == source_release_id,
                Observation.period_start == date(year, 1, 1),
            )
        )
        if observation is None:
            observation = Observation(
                metric_id=metric.id,
                source_release_id=source_release_id,
                provenance_resolved_claim_id=resolved_row.id,
                period_start=date(year, 1, 1),
                period_end=date(year, 12, 31),
                temporal_precision=TemporalPrecision.YEAR,
                temporal_assignment=TemporalAssignment.DIRECT_RECORD,
                date_role=DateRole.REPORTED,
                value_numeric=claim.lower_bound,
                data_status=claim.data_status,
            )
            session.add(observation)
            session.flush()
        coverage = session.scalar(
            select(MetricCoverage).where(
                MetricCoverage.metric_id == metric.id,
                MetricCoverage.source_release_id == source_release_id,
                MetricCoverage.period_start == date(year, 1, 1),
            )
        )
        if coverage is None:
            session.add(
                MetricCoverage(
                    metric_id=metric.id,
                    source_release_id=source_release_id,
                    provenance_resolved_claim_id=resolved_row.id,
                    period_start=date(year, 1, 1),
                    period_end=date(year, 12, 31),
                    coverage_fraction=Decimal("1"),
                    data_status=claim.data_status,
                    comparability_status=ComparabilityStatus.COMPARABLE,
                )
            )
        for selected_date in editorial_dates:
            if year == selected_date.year:
                section = (
                    "typical_day_in_this_year"
                    if claim.claim_type in {"annual_births", "annual_deaths"}
                    else "wider_historical_context"
                )
                record_editorial_selection(
                    session,
                    profile_date=selected_date,
                    section_key=section,
                    resolved_claim_id=resolved_row.id,
                    status=EditorialSelectionStatus.SELECTED,
                    display_rank=resolved_count + 1,
                    rationale=(
                        "Selected official annual context for this profile date."
                    ),
                    reviewed_by="development-fixture-review",
                )
        resolved_count += 1
    for task in session.scalars(
        select(ReviewTask).where(
            ReviewTask.claim_id.in_([claim.id for claim in claims]),
            ReviewTask.status.in_(("open", "in_progress")),
        )
    ):
        task.status = "resolved"
        task.completed_at = datetime.now(UTC)
    for year in sorted(SUPPORTED_YEARS):
        year_claims = {
            claim.claim_type: claim for claim in claims_by_year[year]
        }
        for predicate in ("annual_births", "annual_deaths"):
            claim = year_claims[predicate]
            resolved_input = session.scalar(
                select(ResolvedClaim)
                .where(
                    ResolvedClaim.canonical_key
                    == f"un-wpp:world:{year}:{predicate}"
                )
                .order_by(ResolvedClaim.version.desc())
            )
            assert resolved_input is not None
            annual_thousands = Decimal(
                str((claim.assertion_json or {})["value"])
            )
            days = Decimal(366 if calendar.isleap(year) else 365)
            daily = (annual_thousands * Decimal("1000") / days).quantize(
                Decimal("1"), rounding=ROUND_HALF_UP
            )
            fingerprint = content_hash(
                {
                    "resolved_claim_id": str(resolved_input.id),
                    "resolved_version": resolved_input.version,
                    "days": str(days),
                    "annual_thousands": str(annual_thousands),
                }
            )
            daily_metric_key = f"app:average_daily_{predicate[7:]}"
            daily_metric = session.scalar(
                select(Metric).where(Metric.metric_key == daily_metric_key)
            )
            if daily_metric is None:
                daily_metric = Metric(
                    metric_key=daily_metric_key,
                    display_name=f"Average daily {predicate[7:]}",
                    unit="persons per day",
                    definition=(
                        "Uniform daily equivalent calculated from an annual "
                        "total; not a date-specific observation."
                    ),
                    data_status=DataStatus.MODELED,
                    provenance_resolved_claim_id=resolved_input.id,
                    methodology_id=methodology.id,
                )
                session.add(daily_metric)
                session.flush()
            elif year == BASELINE_YEAR:
                daily_metric.provenance_resolved_claim_id = resolved_input.id
                daily_metric.methodology_id = methodology.id
            derived = session.scalar(
                select(DerivedValue).where(
                    DerivedValue.value_kind
                    == f"average_daily_{predicate[7:]}",
                    DerivedValue.period_start == date(year, 1, 1),
                    DerivedValue.input_fingerprint == fingerprint,
                )
            )
            if derived is None:
                derived = DerivedValue(
                    metric_id=daily_metric.id,
                    methodology_id=methodology.id,
                    provenance_resolved_claim_id=resolved_input.id,
                    value_kind=f"average_daily_{predicate[7:]}",
                    period_start=date(year, 1, 1),
                    period_end=date(year, 12, 31),
                    temporal_assignment=(
                        TemporalAssignment.UNIFORM_PERIOD_ALLOCATION
                    ),
                    value_numeric=daily,
                    value_json={
                        "annual_total_thousands": str(annual_thousands),
                        "days_in_year": int(days),
                        "average_daily_equivalent": int(daily),
                        "display_precision": "nearest whole person",
                        "source_data_status": claim.data_status.value,
                    },
                    data_status=claim.data_status,
                    comparability_status=ComparabilityStatus.COMPARABLE,
                    input_fingerprint=fingerprint,
                    calculation_version="0.4.0",
                )
                session.add(derived)
                session.flush()
                session.add(
                    DerivedValueInput(
                        derived_value_id=derived.id,
                        resolved_claim_id=resolved_input.id,
                        input_role="primary",
                    )
                )
            for selected_date in editorial_dates:
                if selected_date.year == year:
                    record_editorial_selection(
                        session,
                        profile_date=selected_date,
                        section_key="typical_day_in_this_year",
                        derived_value_id=derived.id,
                        status=EditorialSelectionStatus.SELECTED,
                        display_rank=(
                            1 if predicate == "annual_births" else 2
                        ),
                        rationale=(
                            "Selected annual total converted to a uniform daily "
                            "equivalent."
                        ),
                        reviewed_by="development-fixture-review",
                    )
    existing_assessment = session.scalar(
        select(QualityAssessment).where(
            QualityAssessment.source_release_id == source_release_id,
            QualityAssessment.assessment_kind
            == "un_wpp_supported_context_quality_v1",
        )
    )
    if existing_assessment is None:
        session.add(
            QualityAssessment(
            source_release_id=source_release_id,
            methodology_id=methodology.id,
            assessment_kind="un_wpp_supported_context_quality_v1",
            findings={
                "source_resolution": "annual",
                "data_status": "estimated through 2023; modeled projection after",
                "coverage": "World aggregate, every year from 1950 through 2025",
                "daily_equivalent": "uniform allocation, not date-specific",
            },
            public_grade="B",
            public_explanation=(
                "Grade B: official annual UN estimates and medium projections "
                "with clear methodology and complete 1950-2025 coverage; daily "
                "values are transparent uniform equivalents, not date observations."
            ),
            legal_review_status=LegalReviewStatus.NOT_REQUIRED,
            )
        )
    session.flush()
    return resolved_count


def _claim_provenance(
    *,
    claim: Claim,
    resolved: ResolvedClaim,
    release: SourceRelease,
    source: Source,
    methodology: Methodology,
) -> dict[str, object]:
    return {
        "root_type": "resolved_claim",
        "published_statement": "Selected annual UN demographic context.",
        "resolved_claim": {
            "canonical_key": resolved.canonical_key,
            "version": resolved.version,
            "method": resolved.resolution_method.value,
            "rationale": resolved.rationale,
        },
        "supporting_claims": [
            {
                "predicate": claim.claim_type,
                "value": claim.assertion_json,
                "source_record_locator": claim.source_record_locator,
                "source_record_hash_sha256": claim.source_record_hash_sha256,
            }
        ],
        "dissenting_claims": [],
        "source_release": {
            "source": source.name,
            "publisher": source.publisher,
            "release": release.release_label,
            "source_url": release.source_url,
            "raw_checksum_sha256": release.raw_checksum_sha256,
            "retrieved_at": release.retrieved_at.isoformat(),
        },
        "methodology": {
            "name": methodology.name,
            "version": methodology.version,
            "description": methodology.description,
        },
    }


def _derived_provenance(
    *,
    derived: DerivedValue,
    claim: Claim,
    release: SourceRelease,
    source: Source,
    methodology: Methodology,
) -> dict[str, object]:
    return {
        "root_type": "derived_value",
        "published_statement": "Uniform daily equivalent derived from an annual total.",
        "derived_value": {
            "kind": derived.value_kind,
            "calculation_version": derived.calculation_version,
            "value": derived.value_json,
        },
        "supporting_claims": [
            {
                "predicate": claim.claim_type,
                "value": claim.assertion_json,
                "source_record_locator": claim.source_record_locator,
                "source_record_hash_sha256": claim.source_record_hash_sha256,
            }
        ],
        "dissenting_claims": [],
        "source_release": {
            "source": source.name,
            "publisher": source.publisher,
            "release": release.release_label,
            "source_url": release.source_url,
            "raw_checksum_sha256": release.raw_checksum_sha256,
            "retrieved_at": release.retrieved_at.isoformat(),
        },
        "methodology": {
            "name": methodology.name,
            "version": methodology.version,
            "description": methodology.description,
        },
    }


def build_un_wpp_profile_content(
    session: Session,
    *,
    profile_date: date = GOLDEN_DATE,
) -> WPPProfileContent:
    profile_year = profile_date.year
    if profile_year not in SUPPORTED_YEARS:
        raise ValueError("UN WPP context is unavailable before 1950 or after 2025.")
    source = session.scalar(
        select(Source).where(Source.slug == UN_WPP_SOURCE_SLUG)
    )
    if source is None:
        raise ValueError("UN WPP context has not been ingested.")
    release = session.scalar(
        select(SourceRelease)
        .where(SourceRelease.source_id == source.id)
        .order_by(SourceRelease.ingested_at.desc())
    )
    if release is None:
        raise ValueError("UN WPP context has no source release.")
    methodology = _methodology(session)
    claims = {
        claim.claim_type: claim
        for claim in session.scalars(
            select(Claim).where(
                Claim.source_release_id == release.id,
                Claim.temporal_start == date(profile_year, 1, 1),
            )
        )
    }
    required_claims = {
        "population_midyear",
        "annual_births",
        "annual_deaths",
        "life_expectancy",
        "under_five_mortality",
    }
    if set(claims) != required_claims:
        raise ValueError("UN WPP selected-year claims are incomplete.")
    if any(
        claim.assertion_status != ClaimAssertionStatus.ACCEPTED
        for claim in claims.values()
    ):
        raise ValueError("The latest UN WPP release has not completed review.")
    if session.scalar(
        select(ReviewTask.id).where(
            ReviewTask.claim_id.in_([claim.id for claim in claims.values()]),
            ReviewTask.status.in_(("open", "in_progress")),
        )
    ) is not None:
        raise ValueError("The latest UN WPP release has pending review tasks.")
    resolved = {
        predicate: session.scalar(
            select(ResolvedClaim)
            .join(
                ResolvedClaimEvidence,
                ResolvedClaimEvidence.resolved_claim_id == ResolvedClaim.id,
            )
            .where(
                ResolvedClaimEvidence.claim_id == claim.id,
                ResolvedClaimEvidence.stance == "supporting",
            )
            .order_by(ResolvedClaim.version.desc())
        )
        for predicate, claim in claims.items()
    }
    if any(row is None for row in resolved.values()):
        raise ValueError("UN WPP selected claims have not all been resolved.")
    resolved_rows = {
        predicate: row for predicate, row in resolved.items() if row is not None
    }
    derived: dict[str, DerivedValue] = {}
    for row in session.scalars(
        select(DerivedValue)
        .where(
            DerivedValue.methodology_id == methodology.id,
            DerivedValue.period_start == date(profile_year, 1, 1),
        )
        .order_by(DerivedValue.created_at.desc())
    ):
        predicate = {
            "average_daily_births": "annual_births",
            "average_daily_deaths": "annual_deaths",
        }.get(row.value_kind)
        if predicate is None or row.value_kind in derived:
            continue
        input_ids = set(
            session.scalars(
                select(DerivedValueInput.resolved_claim_id).where(
                    DerivedValueInput.derived_value_id == row.id
                )
            )
        )
        if input_ids == {resolved_rows[predicate].id}:
            derived[row.value_kind] = row
    required_derived = {"average_daily_births", "average_daily_deaths"}
    if not required_derived <= set(derived):
        raise ValueError("UN WPP daily equivalents have not been reviewed.")
    assert_release_publication_eligible(
        session,
        source_release_id=release.id,
        profile_date=profile_date,
        resolved_root_ids_by_section={
            "wider_historical_context": {
                resolved_rows[key].id
                for key in (
                    "population_midyear",
                    "life_expectancy",
                    "under_five_mortality",
                )
            }
        },
        derived_root_ids_by_section={
            "typical_day_in_this_year": {
                derived[key].id for key in required_derived
            }
        },
    )
    typical: list[dict[str, object]] = []
    evidence: list[PublicationStatementEvidenceInput] = []
    for index, (predicate, label) in enumerate(
        (("annual_births", "births"), ("annual_deaths", "deaths"))
    ):
        value = derived[f"average_daily_{label}"]
        if value.value_numeric is None:
            raise ValueError("UN WPP daily equivalent is missing its value.")
        daily = int(value.value_numeric)
        status = claims[predicate].data_status.value
        is_projection = claims[predicate].data_status == DataStatus.MODELED
        projection_qualifier = (
            ", based on the UN WPP medium-variant projection"
            if is_projection
            else ""
        )
        selected_date_label = f"{profile_date:%B} {profile_date.day}"
        typical.append(
            {
                "statement_id": f"average-daily-{label}",
                "statement": (
                    f"Average daily {label} in {profile_year}"
                    f"{projection_qualifier}: about {daily:,}. "
                    "This is an average daily equivalent based on the annual total, "
                    f"not an observation for {selected_date_label}."
                ),
                "details": {
                    **(value.value_json or {}),
                    "temporal_assignment": "uniform_period_allocation",
                    "data_status": status,
                    "interpretation": (
                        "Average daily equivalent based on "
                        + (
                            "a medium-variant projected annual total. "
                            if is_projection
                            else "an annual total. "
                        )
                        + "This is not "
                        f"the number observed on {selected_date_label}."
                    ),
                },
                "provenance_note": (
                    "UN WPP "
                    + (
                        "medium-variant projected annual value"
                        if is_projection
                        else "annual value"
                    )
                    + " divided by "
                    f"{366 if calendar.isleap(profile_year) else 365} days; "
                    "not date-specific."
                ),
                "provenance": _derived_provenance(
                    derived=value,
                    claim=claims[predicate],
                    release=release,
                    source=source,
                    methodology=methodology,
                ),
            }
        )
        evidence.append(
            PublicationStatementEvidenceInput(
                statement_path=f"/sections/typical_day_in_this_year/{index}",
                derived_value_id=value.id,
            )
        )
    population_thousands = Decimal(
        str(resolved_rows["population_midyear"].resolved_value["value"])
    )
    life_expectancy = Decimal(
        str(resolved_rows["life_expectancy"].resolved_value["value"])
    )
    under_five_mortality = Decimal(
        str(resolved_rows["under_five_mortality"].resolved_value["value"])
    )
    population_verb = (
        "estimates"
        if claims["population_midyear"].data_status == DataStatus.ESTIMATED
        else "projects"
    )
    life_expectancy_verb = (
        "estimates"
        if claims["life_expectancy"].data_status == DataStatus.ESTIMATED
        else "projects"
    )
    mortality_verb = (
        "estimates"
        if claims["under_five_mortality"].data_status == DataStatus.ESTIMATED
        else "projects"
    )
    context_specs = (
        (
            "population_midyear",
            "world-population",
            (
                f"UN WPP {population_verb} the mid-{profile_year} world "
                "population at about "
                f"{population_thousands / Decimal('1000000'):.3f} billion."
            ),
        ),
        (
            "life_expectancy",
            "world-life-expectancy",
            (
                f"UN WPP {life_expectancy_verb} global life expectancy at birth "
                f"in {profile_year} at "
                f"{life_expectancy:.2f} years."
            ),
        ),
        (
            "under_five_mortality",
            "world-under-five-mortality",
            (
                f"UN WPP {mortality_verb} {profile_year} global under-five "
                "mortality at about "
                f"{under_five_mortality:.0f} deaths per 1,000 live births."
            ),
        ),
    )
    context: list[dict[str, object]] = []
    for index, (predicate, statement_id, statement) in enumerate(context_specs):
        claim = claims[predicate]
        resolved_row = resolved_rows[predicate]
        context.append(
            {
                "statement_id": statement_id,
                "statement": statement,
                "details": {
                    **(claim.assertion_json or {}),
                    "temporal_assignment": "direct_record",
                    "temporal_precision": "year",
                    "data_status": claim.data_status.value,
                },
                "provenance_note": (
                    "Official annual UN WPP estimate."
                    if claim.data_status == DataStatus.ESTIMATED
                    else "Official annual UN WPP medium projection."
                ),
                "provenance": _claim_provenance(
                    claim=claim,
                    resolved=resolved_row,
                    release=release,
                    source=source,
                    methodology=methodology,
                ),
            }
        )
        evidence.append(
            PublicationStatementEvidenceInput(
                statement_path=f"/sections/wider_historical_context/{index}",
                resolved_claim_id=resolved_row.id,
            )
        )
    return WPPProfileContent(
        typical_statements=typical,
        context_statements=context,
        evidence=evidence,
        source_release_id=release.id,
        methodology=methodology,
        resolved_claims=tuple(
            sorted(resolved_rows.values(), key=lambda row: row.canonical_key)
        ),
    )
