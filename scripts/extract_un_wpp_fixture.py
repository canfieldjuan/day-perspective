from __future__ import annotations

import argparse
import csv
import hashlib
from pathlib import Path

from openpyxl import load_workbook

WORKBOOK_SHA256 = (
    "98e34d9b65b53858cd08a57a566e45050b08093ad85ba5714fe6fbd78055ae6d"
)
FIELDS = (
    "location",
    "location_code",
    "year",
    "variant",
    "population_july_thousands",
    "births_thousands",
    "deaths_thousands",
    "life_expectancy_years",
    "under_five_mortality_per_1000",
)
COLUMNS = {
    "variant": 1,
    "location": 2,
    "location_code": 4,
    "year": 10,
    "population_july_thousands": 12,
    "births_thousands": 23,
    "deaths_thousands": 30,
    "life_expectancy_years": 34,
    "under_five_mortality_per_1000": 50,
}


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Extract the public World 1950-2025 WPP fixture."
    )
    parser.add_argument("workbook", type=Path)
    parser.add_argument("output", type=Path)
    args = parser.parse_args()

    payload = args.workbook.read_bytes()
    checksum = hashlib.sha256(payload).hexdigest()
    if checksum != WORKBOOK_SHA256:
        raise ValueError(
            f"Expected pinned WPP workbook {WORKBOOK_SHA256}, received {checksum}."
        )

    workbook = load_workbook(
        filename=args.workbook,
        read_only=True,
        data_only=True,
    )
    rows: list[dict[str, object]] = []
    for sheet_name, years in (
        ("Estimates", range(1950, 2024)),
        ("Medium variant", range(2024, 2026)),
    ):
        sheet = workbook[sheet_name]
        for row in sheet.iter_rows(min_row=18, values_only=True):
            if row[COLUMNS["location_code"]] != 900:
                if rows and rows[-1]["year"] in years:
                    break
                continue
            year = int(row[COLUMNS["year"]])
            if year not in years:
                continue
            rows.append({field: row[index] for field, index in COLUMNS.items()})
    workbook.close()

    if [int(row["year"]) for row in rows] != list(range(1950, 2026)):
        raise ValueError("The workbook did not yield exactly World years 1950-2025.")
    args.output.parent.mkdir(parents=True, exist_ok=True)
    with args.output.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=FIELDS)
        writer.writeheader()
        writer.writerows(rows)


if __name__ == "__main__":
    main()
